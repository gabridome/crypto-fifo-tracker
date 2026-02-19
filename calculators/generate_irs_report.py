"""
Generate Portuguese IRS Report for Crypto Capital Gains (FIFO)

FORMAT SOURCE:
  Autoridade Tributária e Aduaneira — MOD 3 IRS Anexo G1
  Quadro 7: Criptoativos que não constituam valores mobiliários
            detidos por período superior ou igual a 365 dias
  URL: https://www.lexpoint.pt/Fileget.aspx?FileId=52911
  Retrieved: 2026-02-08

  Anexo J — Quadro 9.4A: Alienação onerosa de criptoativos
            detidos por período inferior a 365 dias (plataformas estrangeiras)
  URL: https://www.lexpoint.pt/Fileget.aspx?FileId=52914
  Retrieved: 2026-02-08

LEGAL BASIS:
  Art.º 10.º, n.º 1, al. k), n.º 19 e n.º 22 do Código do IRS
  Gains from crypto held ≥365 days: exempt (Anexo G1, Quadro 7)
  Gains from crypto held <365 days via foreign platforms: taxed 28% (Anexo J, Q.9.4A)
  FIFO method required (art.º 10.º CIRS)

Usage:
    python3 generate_irs_report.py <year> [db_path]

Examples:
    python3 generate_irs_report.py 2024
    python3 generate_irs_report.py 2025 /path/to/data/crypto_fifo.db

Output:
    data/reports/IRS_Crypto_FIFO_<year>.xlsx
"""

import sqlite3
import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = 'data/crypto_fifo.db'

# --- Exchange -> (Country Code per Tabela X Anexo J, Country Name) ---
EXCHANGE_COUNTRY = {
    'Coinbase':        ('840', 'Estados Unidos'),
    'Coinbase Prime':  ('840', 'Estados Unidos'),
    'Binance':         ('136', 'Ilhas Caimao'),
    'Binance Card':    ('136', 'Ilhas Caimao'),
    'Binance OTC':     ('136', 'Ilhas Caimao'),
    'Bitstamp':        ('826', 'Reino Unido'),
    'Kraken':          ('840', 'Estados Unidos'),
    'Bitfinex':        ('092', 'Ilhas Virgens Britanicas'),
    'Mt.Gox':          ('392', 'Japao'),
    'TRT':             ('380', 'Italia'),
    'Wirex':           ('826', 'Reino Unido'),
    'Revolut':         ('440', 'Lituania'),
    'GDTRE':           ('380', 'Italia'),
    'Coinpal':         ('840', 'Estados Unidos'),
}

DOMESTIC_EXCHANGES = set()  # Portuguese platforms with NIF

OWNER_TITULAR = 'A'  # Sujeito Passivo A

# --- Styles ---
HEADER_FONT = Font(name='Arial', bold=True, size=9, color='000000')
HEADER_FILL = PatternFill('solid', fgColor='B4C6E7')
TITLE_FONT = Font(name='Arial', bold=True, size=12, color='000000')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=9, color='333333')
DATA_FONT = Font(name='Arial', size=9)
BOLD_FONT = Font(name='Arial', bold=True, size=9)
MONEY_FMT = '#,##0.00'
THIN_BORDER = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin')
)
SUMMARY_FILL = PatternFill('solid', fgColor='D9E2F3')
EXEMPT_FILL = PatternFill('solid', fgColor='E2EFDA')
TAXABLE_FILL = PatternFill('solid', fgColor='FCE4EC')
SOURCE_FONT = Font(name='Arial', size=7, italic=True, color='666666')


def get_daily_sales(db_path, year):
    """
    Aggregate sales by DAY + EXCHANGE + TAX STATUS as required by Portuguese AT.
    Each row = one day on one exchange, separated by exempt/taxable.
    This is critical: exempt (>=365d) and taxable (<365d) MUST be separate rows
    because they go to different annexes (G1 vs J).
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    query = """
        SELECT
            date(slm.sale_date) as sale_day,
            t.exchange_name,
            GROUP_CONCAT(DISTINCT slm.cryptocurrency) as cryptocurrencies,
            MIN(slm.purchase_date) as earliest_purchase,
            SUM(slm.amount_sold) as amount,
            SUM(slm.proceeds) as proceeds,
            SUM(slm.cost_basis) as cost_basis,
            SUM(slm.gain_loss) as gain_loss,
            SUM(slm.cost_basis - (slm.amount_sold * slm.purchase_price_per_unit)) as purchase_fees,
            MIN(slm.holding_period_days) as min_holding_days,
            (slm.holding_period_days >= 365) as is_exempt,
            COUNT(*) as num_transactions
        FROM sale_lot_matches slm
        JOIN transactions t ON slm.sale_transaction_id = t.id
        WHERE slm.sale_date >= ? AND slm.sale_date < ?
        GROUP BY date(slm.sale_date), t.exchange_name, (slm.holding_period_days >= 365)
        ORDER BY sale_day ASC, t.exchange_name ASC
    """
    rows = conn.execute(query, (f'{year}-01-01', f'{year+1}-01-01')).fetchall()

    # Get sale fees per day per exchange per tax status
    # We need to split fees proportionally between exempt and taxable
    fee_query = """
        SELECT
            date(slm.sale_date) as sale_day,
            t.exchange_name,
            (slm.holding_period_days >= 365) as is_exempt,
            SUM(t.fee_amount * slm.amount_sold / t.amount) as proportional_sale_fees
        FROM sale_lot_matches slm
        JOIN transactions t ON slm.sale_transaction_id = t.id
        WHERE slm.sale_date >= ? AND slm.sale_date < ?
          AND t.fee_amount > 0
        GROUP BY date(slm.sale_date), t.exchange_name, (slm.holding_period_days >= 365)
    """
    fee_rows = conn.execute(fee_query, (f'{year}-01-01', f'{year+1}-01-01')).fetchall()
    sale_fees = {(r['sale_day'], r['exchange_name'], r['is_exempt']): r['proportional_sale_fees']
                 for r in fee_rows}

    conn.close()

    result = []
    for r in rows:
        d = dict(r)
        key = (d['sale_day'], d['exchange_name'], d['is_exempt'])
        d['sale_fees'] = sale_fees.get(key, 0) or 0
        result.append(d)
    return result


def classify_day(day_row):
    exempt = day_row['min_holding_days'] >= 365
    exchange = day_row['exchange_name'] or 'Unknown'
    domestic = exchange in DOMESTIC_EXCHANGES
    country_code, country_name = EXCHANGE_COUNTRY.get(exchange, ('999', 'Desconhecido'))

    if exempt:
        anexo = 'Anexo G1 Quadro 07'
    elif not domestic:
        anexo = 'Anexo J Quadro 9.4A'
    else:
        anexo = 'Anexo G Quadro 18A'

    return {**day_row, 'exempt': exempt, 'domestic': domestic,
            'anexo': anexo, 'country_code': country_code, 'country_name': country_name}


def parse_date_parts(iso_str):
    """Parse ISO date -> (ano, mes, dia) as integers"""
    if not iso_str:
        return (None, None, None)
    try:
        dt = datetime.fromisoformat(iso_str)
        return (dt.year, dt.month, dt.day)
    except ValueError:
        parts = iso_str[:10].split('-')
        return (int(parts[0]), int(parts[1]), int(parts[2]))


def apply_border(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


# =====================================================================
# SHEET 1: Anexo G1 — Quadro 7 (exempt, >=365 days)
# =====================================================================
def write_quadro7(ws, year, sales):
    ws.title = 'Anexo G1 - Quadro 7'
    ws.sheet_properties.tabColor = '548235'
    exempt = [s for s in sales if s['exempt']]

    # Source
    ws.merge_cells('A1:N1')
    ws['A1'] = ('Formato: Autoridade Tributaria — MOD 3 IRS Anexo G1 | '
                'URL: https://www.lexpoint.pt/Fileget.aspx?FileId=52911 | '
                'Consultado: 2026-02-08')
    ws['A1'].font = SOURCE_FONT

    # Title
    ws.merge_cells('A3:N3')
    ws['A3'] = ('QUADRO 7 — CRIPTOATIVOS QUE NAO CONSTITUAM VALORES MOBILIARIOS '
                'DETIDOS POR PERIODO >= 365 DIAS')
    ws['A3'].font = TITLE_FONT

    ws.merge_cells('A4:N4')
    ws['A4'] = (f'Alienacao onerosa [art. 10., n. 1, al. k) e n. 19, do CIRS] '
                f'/ Perda da qualidade de residente — Ano: {year}')
    ws['A4'].font = SUBTITLE_FONT

    # Header row 1 (merged groups)
    r1 = 6
    r2 = 7

    # Col A: TITULAR (merged 2 rows)
    ws.merge_cells(f'A{r1}:A{r2}')
    c = ws[f'A{r1}']
    c.value = 'TITULAR'
    c.font = HEADER_FONT; c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Cols B-C: ENTIDADE GESTORA
    ws.merge_cells(f'B{r1}:C{r1}')
    c = ws[f'B{r1}']
    c.value = 'ENTIDADE GESTORA'
    c.font = HEADER_FONT; c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)

    # Cols D-G: REALIZACAO
    ws.merge_cells(f'D{r1}:G{r1}')
    c = ws[f'D{r1}']
    c.value = 'REALIZACAO'
    c.font = HEADER_FONT; c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)

    # Cols H-K: AQUISICAO
    ws.merge_cells(f'H{r1}:K{r1}')
    c = ws[f'H{r1}']
    c.value = 'AQUISICAO'
    c.font = HEADER_FONT; c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)

    # Col L: Despesas e encargos (merged 2 rows)
    ws.merge_cells(f'L{r1}:L{r2}')
    c = ws[f'L{r1}']
    c.value = 'Despesas e\nencargos'
    c.font = HEADER_FONT; c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Col M: Pais da contraparte (merged 2 rows)
    ws.merge_cells(f'M{r1}:M{r2}')
    c = ws[f'M{r1}']
    c.value = 'Pais da\ncontraparte'
    c.font = HEADER_FONT; c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Col N: Exchange (internal note, not in official form)
    ws.merge_cells(f'N{r1}:N{r2}')
    c = ws[f'N{r1}']
    c.value = 'Exchange\n(nota interna)'
    c.font = Font(name='Arial', bold=True, size=8, italic=True, color='666666')
    c.fill = PatternFill('solid', fgColor='E0E0E0')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Header row 2 (sub-columns)
    subs = {'B': 'NIF\nPortugues', 'C': 'Pais',
            'D': 'Ano', 'E': 'Mes', 'F': 'Dia', 'G': 'Valor',
            'H': 'Ano', 'I': 'Mes', 'J': 'Dia', 'K': 'Valor'}
    for col_letter, label in subs.items():
        c = ws[f'{col_letter}{r2}']
        c.value = label
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    apply_border(ws, r1, r2, 1, 14)

    # Data
    data_start = 8
    row = data_start
    for s in exempt:
        sy, sm, sd = parse_date_parts(s['sale_day'])
        ay, am, ad = parse_date_parts(s['earliest_purchase'])
        code = s['country_code']
        exchange = s['exchange_name'] or ''

        sale_fee = float(s.get('sale_fees') or 0)
        purch_fee = float(s.get('purchase_fees') or 0)
        despesas = sale_fee + purch_fee

        # Official form: Valor = gross amounts; fees go in Despesas separately
        valor_real = s['proceeds'] + sale_fee
        valor_aquis = s['cost_basis'] - purch_fee

        data = [
            (1, OWNER_TITULAR), (2, ''), (3, code),
            (4, sy), (5, sm), (6, sd), (7, valor_real),
            (8, ay), (9, am), (10, ad), (11, valor_aquis),
            (12, despesas), (13, code), (14, exchange),
        ]
        for col, val in data:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT; cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            if col in (7, 11, 12):
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal='right')
        row += 1

    # SOMA
    if exempt:
        last = row - 1
        ws.merge_cells(f'A{row}:F{row}')
        c = ws[f'A{row}']
        c.value = 'SOMA'; c.font = BOLD_FONT; c.fill = SUMMARY_FILL
        c.alignment = Alignment(horizontal='right')
        for col in (7, 11, 12):
            cell = ws.cell(row=row, column=col)
            cl = get_column_letter(col)
            cell.value = f'=SUM({cl}{data_start}:{cl}{last})'
            cell.font = BOLD_FONT; cell.fill = SUMMARY_FILL
            cell.number_format = MONEY_FMT; cell.alignment = Alignment(horizontal='right')
        apply_border(ws, row, row, 1, 14)

    widths = {'A':8, 'B':12, 'C':8, 'D':6, 'E':5, 'F':5, 'G':14,
              'H':6, 'I':5, 'J':5, 'K':14, 'L':14, 'M':10, 'N':16}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v


# =====================================================================
# SHEET 2: Anexo J — Quadro 9.4A (taxable, <365 days, foreign)
# =====================================================================
def write_quadro94a(ws, year, sales):
    ws.title = 'Anexo J - Quadro 9.4A'
    ws.sheet_properties.tabColor = 'C00000'
    taxable = [s for s in sales if not s['exempt'] and not s['domestic']]

    ws.merge_cells('A1:P1')
    ws['A1'] = ('Formato: Autoridade Tributaria — MOD 3 IRS Anexo J | '
                'URL: https://www.lexpoint.pt/Fileget.aspx?FileId=52914 | '
                'Consultado: 2026-02-08')
    ws['A1'].font = SOURCE_FONT

    ws.merge_cells('A3:P3')
    ws['A3'] = ('QUADRO 9.4A — ALIENACAO ONEROSA DE CRIPTOATIVOS (< 365 DIAS) — '
                'PLATAFORMAS ESTRANGEIRAS')
    ws['A3'].font = TITLE_FONT

    ws.merge_cells('A4:P4')
    ws['A4'] = f'Art. 10., n. 1, al. k), n. 19 e n. 22, do CIRS — Ano: {year}'
    ws['A4'].font = SUBTITLE_FONT

    fill_h = PatternFill('solid', fgColor='E6B8B7')
    r1, r2 = 6, 7

    groups = [
        ('A', 'A', 'TITULAR', True),
        ('B', 'C', 'ENTIDADE GESTORA', False),
        ('D', 'D', 'Pais da\nFonte', True),
        ('E', 'H', 'REALIZACAO', False),
        ('I', 'L', 'AQUISICAO', False),
        ('M', 'M', 'Despesas e\nencargos', True),
        ('N', 'N', 'Imposto pago\nno estrangeiro', True),
        ('O', 'O', 'Pais da\ncontraparte', True),
        ('P', 'P', 'Exchange\n(nota interna)', True),
    ]
    for start, end, label, merge_rows in groups:
        if start != end:
            ws.merge_cells(f'{start}{r1}:{end}{r1}')
        elif merge_rows:
            ws.merge_cells(f'{start}{r1}:{start}{r2}')
        c = ws[f'{start}{r1}']
        c.value = label; c.font = HEADER_FONT; c.fill = fill_h
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    subs = {'B': 'NIF\nPortugues', 'C': 'Pais',
            'E': 'Ano', 'F': 'Mes', 'G': 'Dia', 'H': 'Valor',
            'I': 'Ano', 'J': 'Mes', 'K': 'Dia', 'L': 'Valor'}
    for k, v in subs.items():
        c = ws[f'{k}{r2}']
        c.value = v; c.font = HEADER_FONT; c.fill = fill_h
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    apply_border(ws, r1, r2, 1, 16)

    data_start = 8
    row = data_start

    if not taxable:
        ws.merge_cells(f'A{row}:P{row}')
        c = ws[f'A{row}']
        c.value = 'Nenhuma alienacao tributavel de criptoativos neste ano.'
        c.font = Font(name='Arial', size=10, italic=True, color='666666')
        c.alignment = Alignment(horizontal='center')
    else:
        for s in taxable:
            sy, sm, sd = parse_date_parts(s['sale_day'])
            ay, am, ad = parse_date_parts(s['earliest_purchase'])
            code = s['country_code']
            sf = float(s.get('sale_fees') or 0)
            pf = float(s.get('purchase_fees') or 0)
            data = [
                (1, OWNER_TITULAR), (2, ''), (3, code), (4, code),
                (5, sy), (6, sm), (7, sd), (8, s['proceeds'] + sf),
                (9, ay), (10, am), (11, ad), (12, s['cost_basis'] - pf),
                (13, sf + pf), (14, 0), (15, code), (16, s['exchange_name'] or ''),
            ]
            for col, val in data:
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = DATA_FONT; cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center')
                if col in (8, 12, 13, 14):
                    cell.number_format = MONEY_FMT
                    cell.alignment = Alignment(horizontal='right')
            row += 1

        last = row - 1
        ws.merge_cells(f'A{row}:G{row}')
        c = ws[f'A{row}']
        c.value = 'SOMA'; c.font = BOLD_FONT; c.fill = SUMMARY_FILL
        c.alignment = Alignment(horizontal='right')
        for col in (8, 12, 13):
            cell = ws.cell(row=row, column=col)
            cl = get_column_letter(col)
            cell.value = f'=SUM({cl}{data_start}:{cl}{last})'
            cell.font = BOLD_FONT; cell.fill = SUMMARY_FILL; cell.number_format = MONEY_FMT
        apply_border(ws, row, row, 1, 16)

    widths = {'A':8, 'B':12, 'C':8, 'D':8, 'E':6, 'F':5, 'G':5, 'H':14,
              'I':6, 'J':5, 'K':5, 'L':14, 'M':14, 'N':14, 'O':10, 'P':16}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v


# =====================================================================
# SHEET 3: Resumo (internal summary)
# =====================================================================
def write_summary(ws, year, sales):
    ws.title = f'Resumo {year}'
    ws.sheet_properties.tabColor = '2F5496'
    exempt = [s for s in sales if s['exempt']]
    taxable = [s for s in sales if not s['exempt']]

    ws.merge_cells('A1:H1')
    ws['A1'] = f'RIEPILOGO FISCALE CRIPTO — IRS {year}'
    ws['A1'].font = TITLE_FONT

    ws.merge_cells('A2:H2')
    ws['A2'] = f'Generato: {datetime.now().strftime("%Y-%m-%d %H:%M")} | Metodo FIFO | EUR'
    ws['A2'].font = SOURCE_FONT

    headers = ['Categoria', 'N.', 'BTC', 'Ricavi', 'Costo Base',
               'Plus/Minus', 'Imposta', 'Anexo']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2F5496')
        c.border = THIN_BORDER; c.alignment = Alignment(horizontal='center', wrap_text=True)

    def cat_row(r, label, items, rate, anexo, fill):
        vals = [label, len(items),
                sum(s['amount'] for s in items),
                sum(s['proceeds'] for s in items),
                sum(s['cost_basis'] for s in items),
                sum(s['gain_loss'] for s in items),
                sum(s['gain_loss'] for s in items) * rate, anexo]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = DATA_FONT; c.fill = fill; c.border = THIN_BORDER
            if col == 3: c.number_format = '#,##0.00000000'
            elif col in (4,5,6,7): c.number_format = MONEY_FMT

    cat_row(5, 'Esenti (>=365 gg)', exempt, 0, 'G1 Q.07', EXEMPT_FILL)
    cat_row(6, 'Tassabili (<365 gg)', taxable, 0.28, 'J Q.9.4A', TAXABLE_FILL)

    tot = ['TOTALE', len(sales), sum(s['amount'] for s in sales),
           sum(s['proceeds'] for s in sales), sum(s['cost_basis'] for s in sales),
           sum(s['gain_loss'] for s in sales),
           sum(s['gain_loss'] for s in taxable) * 0.28 if taxable else 0, '']
    for col, val in enumerate(tot, 1):
        c = ws.cell(row=7, column=col, value=val)
        c.font = BOLD_FONT; c.fill = SUMMARY_FILL; c.border = THIN_BORDER
        if col == 3: c.number_format = '#,##0.00000000'
        elif col in (4,5,6,7): c.number_format = MONEY_FMT

    # Exchange breakdown
    ws.cell(row=9, column=1, value='Per Exchange').font = BOLD_FONT
    ex_h = ['Exchange', 'Paese', 'Cod.', 'N.', 'BTC', 'Ricavi', 'Plus/Minus']
    for col, h in enumerate(ex_h, 1):
        c = ws.cell(row=10, column=col, value=h)
        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2F5496'); c.border = THIN_BORDER

    exch = {}
    for s in sales:
        ex = s['exchange_name'] or 'Unknown'
        if ex not in exch: exch[ex] = {'n': 0, 'a': 0, 'p': 0, 'g': 0}
        exch[ex]['n'] += 1; exch[ex]['a'] += s['amount']
        exch[ex]['p'] += s['proceeds']; exch[ex]['g'] += s['gain_loss']

    row = 11
    for name, d in sorted(exch.items()):
        code, country = EXCHANGE_COUNTRY.get(name, ('999', '?'))
        for col, val in enumerate([name, country, code, d['n'], d['a'], d['p'], d['g']], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = DATA_FONT; c.border = THIN_BORDER
            if col == 5: c.number_format = '#,##0.00000000'
            elif col in (6,7): c.number_format = MONEY_FMT
        row += 1

    row += 1
    notes = [
        f'Metodo FIFO (art. 10. CIRS). Periodo: 01/01/{year}-31/12/{year}.',
        f'Dichiarazione: 01/04/{year+1}-30/06/{year+1}. Pagamento: 31/08/{year+1}.',
        'Cripto >=365gg esenti ma da dichiarare (Anexo G1 Q.07).',
        'Cripto <365gg via piattaforme estere: Anexo J Q.9.4A, taxa 28%.',
    ]
    for n in notes:
        ws.cell(row=row, column=1, value=f'  {n}').font = Font(name='Arial', size=8, color='444444')
        row += 1

    for i, w in enumerate([24, 16, 8, 8, 16, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =====================================================================
# SHEET 4: Dettaglio completo (archive)
# =====================================================================
def write_detail(ws, year, sales):
    ws.title = 'Dettaglio'
    ws.sheet_properties.tabColor = '7F7F7F'

    ws.merge_cells('A1:P1')
    ws['A1'] = f'DETTAGLIO AGGREGATI GIORNALIERI CRIPTO {year}'
    ws['A1'].font = TITLE_FONT

    headers = ['N.', 'Cripto', 'Exchange', 'N.Op.',
               'Vend.Ano', 'Vend.Mes', 'Vend.Dia',
               'Acq.Ano', 'Acq.Mes', 'Acq.Dia',
               'Quantita', 'Val.Realizacao', 'Val.Aquisicao',
               'Despesas', 'Giorni Min', 'Anexo']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name='Arial', bold=True, size=8, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='595959')
        c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = THIN_BORDER

    row = 4
    for idx, s in enumerate(sales, 1):
        sy, sm, sd = parse_date_parts(s['sale_day'])
        ay, am, ad = parse_date_parts(s['earliest_purchase'])
        sf = float(s.get('sale_fees') or 0)
        pf = float(s.get('purchase_fees') or 0)

        vals = [idx, s.get('cryptocurrencies', ''), s['exchange_name'], s.get('num_transactions', 1),
                sy, sm, sd, ay, am, ad,
                s['amount'], s['proceeds'] + sf, s['cost_basis'] - pf,
                sf + pf, s['min_holding_days'], s['anexo']]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=8); c.border = THIN_BORDER
            if col == 11: c.number_format = '#,##0.00000000'
            elif col in (12, 13, 14): c.number_format = MONEY_FMT
            elif col == 15: c.number_format = '#,##0'
        fill = EXEMPT_FILL if s['exempt'] else TAXABLE_FILL
        for col in range(1, 17):
            ws.cell(row=row, column=col).fill = fill
        row += 1

    for i, w in enumerate([5, 10, 14, 6, 6, 5, 5, 6, 5, 5, 16, 14, 14, 12, 10, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =====================================================================
# MAIN
# =====================================================================
def main():
    if len(sys.argv) < 2:
        print(f'Usage: python3 {sys.argv[0]} <year> [db_path]')
        sys.exit(1)

    year = int(sys.argv[1])
    db_path = sys.argv[2] if len(sys.argv) > 2 else DB_PATH

    if not os.path.exists(db_path):
        print(f'x Database not found: {db_path}')
        sys.exit(1)

    print('=' * 70)
    print(f'  CRYPTO FIFO — REPORT IRS PORTOGALLO {year}')
    print(f'  Aggregazione: per GIORNO per EXCHANGE')
    print('=' * 70)
    print(f'\n  Database: {os.path.abspath(db_path)}')

    raw_days = get_daily_sales(db_path, year)
    if not raw_days:
        print(f'\n  Nessuna vendita trovata per il {year}.')
        sys.exit(0)

    sales = [classify_day(d) for d in raw_days]
    exempt = [s for s in sales if s['exempt']]
    taxable = [s for s in sales if not s['exempt']]

    total_ops = sum(s.get('num_transactions', 1) for s in sales)
    print(f'\n  Operazioni totali: {total_ops}')
    print(f'  Righe report (aggregati giornalieri): {len(sales)}')
    print(f'    Esenti (>=365gg):  {len(exempt):>4}  |  EUR {sum(s["gain_loss"] for s in exempt):>12,.2f}')
    print(f'    Tassabili (<365):  {len(taxable):>4}  |  EUR {sum(s["gain_loss"] for s in taxable):>12,.2f}')
    print(f'  Exchange: {", ".join(sorted(set(s["exchange_name"] for s in sales)))}')

    wb = Workbook()
    write_quadro7(wb.active, year, sales)
    write_quadro94a(wb.create_sheet(), year, sales)
    write_summary(wb.create_sheet(), year, sales)
    write_detail(wb.create_sheet(), year, sales)

    project_dir = os.path.dirname(os.path.abspath(db_path))
    output_dir = os.path.join(project_dir, 'reports')
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f'IRS_Crypto_FIFO_{year}.xlsx')
    wb.save(output_file)
    print(f'\n  Report: {output_file}')

    tax = sum(s['gain_loss'] for s in taxable) * 0.28 if taxable else 0
    print(f'\n  {"=" * 50}')
    print(f'  Plus-valenza totale:   EUR {sum(s["gain_loss"] for s in sales):>12,.2f}')
    print(f'  Imposta stimata:       EUR {tax:>12,.2f}')
    print(f'  {"=" * 50}')
    if tax == 0:
        print(f'  Nessuna imposta. Dichiarare in Anexo G1 Quadro 07.')
    print('=' * 70)


if __name__ == '__main__':
    main()
